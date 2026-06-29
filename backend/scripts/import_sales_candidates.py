"""
Import high-confidence candidate products from the May 2026 sales report.

- Only keyword-rule (high-confidence) classifications.
- Excludes products that likely already exist in the catalogue.
- price derived from Sales Amount / Qty Sold (rounded).
- Inserted as INACTIVE drafts (is_active=False, in_stock=False, stock_quantity=0).

Usage:
  python scripts/import_sales_candidates.py            # dry run (no writes)
  python scripts/import_sales_candidates.py --commit   # actually insert
"""
import asyncio, sys, re, math
from pathlib import Path
from datetime import datetime, timezone
import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))
from database import products_collection

XLSX = r"C:\Users\alfre\Downloads\Margin_Analysis_May2026.xlsx"

STOP = set('THE A AND OF FOR WITH IN ML L G KG GM RS PCS PC PKT PACK PVC S J CAN REF LT LTR JAR BA NO X FULL NEW BIG SMALL'.split())
def toks(s):
    s = re.sub(r'[^A-Z ]', ' ', s.upper())
    return [t for t in s.split() if len(t) > 2 and t not in STOP]

RULES = [
 (r'LAMP OIL|DHEEPAM|DEEPAM|CAMPHOR|KARPOORAM|AGARBATH|AGARBATTI|SAMBRANI|VIBUTHI|VIBHUTHI|KUNGUM|VERMILION|POOJA|INCENSE|\bTHIRI\b|\bWICK\b|HOMAM|NAMAM|SANDAL STICK', 'Pooja & Religious'),
 (r'NAIL POLISH|SHAMPOO|\bSOAP\b|TOOTH ?PASTE|TOOTH ?BRUSH|HAIR OIL|FACE WASH|FACE CREAM|FAIRNESS|\bFAIR\b|LOTION|\bTALC\b|\bDEO\b|DEODOR|PERFUME|\bRAZOR\b|GILLET|\bGIL\b|SHAVE|SANITARY|STAYFREE|WHISPER|DIAPER|LIPSTICK|KAJAL|MEHND|HENNA|VICCO|PONDS|NIVEA|LAKME|\bCOMB\b|HANDWASH|HAND WASH|COLD CREAM|MOISTURIS|VATIKA|CINTHOL|SANTOOR|LIFEBUOY|MEERA|\bDOVE\b|\bAXE\b|CLOSE UP|COLGATE|\bCOL PASTE\b|H&S|HEAD.*SHOULDER|CLINIC PLUS|SUNSILK|PARACHUTE|\bFEMI\b|TEDIBAR|VASELINE|GARNIER', 'Personal Care'),
 (r'\bBABY\b|PAMPER|HUGGIES|CERELAC|LACTOGEN|NAN PRO|JOHNSON', 'Baby Care'),
 (r'BAND ?AID|DETTOL|SAVLON|SANITIZER|\bBALM\b|VICKS|ZANDU|TABLET|SYRUP|\bMOOV\b|VOLINI|ITCH GUARD|OINTMENT|\bCOTTON\b|THERMO|\bMASK\b|CREP BANDAGE|CREPE', 'Healthcare'),
 (r'DETERGENT|\bSURF\b|ARIEL|\bRIN\b|\bTIDE\b|WHEEL|HENKO|UJALA|PHENYL|PHENYLE|HARPIC|LIZOL|BLEACH|\bCOLIN\b|BROOM|\bMOP\b|SCRUB|TISSUE|NAPKIN|GARBAGE|DUST ?BIN|FAB LIQUID|\bVIM\b|\bEXO\b|\bPRIL\b|TOILET|\bFLOOR\b|MATCHES|\bCANDLE\b|MOSQUITO|GOOD KNIGHT|MORTEIN|ALL OUT|NAPHTHAL|FRESHNER|FRESHENER|AIR FRESH|SCOTCH BRITE|OZONE|PITAMBARI|KOLAPODI|MONKEY 555|\bMAT-', 'Household'),
 (r'\bPEN\b|PENCIL|NOTE ?BOOK|\bNOTES\b|ERASER|SHARPEN|\bCELLO\b|GEOMETRY|GIFT PAPER|\bRIBBON\b|PAPER CLIP|STAPLE|FEVI|SKETCH|CRAYON|COLOUR BOX|GOLD CAMEL|\bSCALE\b|\bTOY\b|BALLOON|MARKER|\bCHART\b|\bBOOKS\b|APPLE TREE', 'Toys & Stationery'),
 (r'NOODLE|\bPASTA\b|MACARONI|MACRONI|VERMICELL|SEMIYA|TOP RAMEN|\bMAGGI\b|YIPPEE', 'Pasta & Noodles'),
 (r'\bDHALL?\b|\bDAL\b|THUVARAM|\bTOOR\b|\bMOONG\b|\bORID\b|\bURAD\b|\bURID\b|\bGRAM\b|CHANA|RAJMA|MASOOR|GREEN GRAM|FRIED GRAM|BLACK GRAM|\bKOLLU\b|HORSE GRAM|\bPEAS\b|KADALAI', 'Pulses & Grains'),
 (r'\bRICE\b|\bPONNI\b|BASMATI|\bSONA\b|\bRAVA\b|\bATTA\b|\bMAIDA\b|\bSOOJI\b|\bWHEAT\b|\bPOHA\b|\bAVAL\b|\bFLOUR\b|\bSEVAI\b|\bGOPURAM\b|\bBULLET\b|\bGTS\b|\bDSM\b|\bCBR\b|\bRAGI\b|MILLET', 'Rice & Cereals'),
 (r'GROUNDNUT OIL|SUNFLOWER OIL|GINGELLY|GIN OIL|\bGN OIL\b|G ?NUT OIL|PALM OIL|PALM\b|\bS\.?F OIL\b|\bSF OIL\b|GOLD WINNER|MR\.?GOLD|IDHAYAM|FORTUNE|RUCHI GOLD|COCONUT OIL|CASTOR OIL|MANTRA OIL|VANASPATI|DALDA|COOKING OIL|REFINED OIL|MAHARAJA G', 'Cooking Oils'),
 (r'\bGHEE\b|PANEER|\bCURD\b|\bMILK\b|BUTTER|CHEESE|KHOVA|\bKOVA\b|\bDAHI\b|LASSI|NANDINI|MILKY MIST|\bAAVIN\b|AAHHAA|FRESH CREAM|\bDAIRY\b', 'Dairy & Fats'),
 (r'\bTEA\b|COFFEE|3 ?ROSES|RED LABEL|TAJ MAHAL|\bBRU\b|NESCAFE|GREEN TEA|TOP STAR|\bBOOST\b|HORLICKS|BOURNVITA|COMPLAN', 'Tea & Coffee'),
 (r'CASHEW|BADAM|ALMOND|PISTA|WALNUT|ANJEER|RAISIN|KISMIS|DRY DATES|\bDATES\b|GROUNDNUT SEED|\bPEANUT\b|DRY FRUIT|\bFIG\b', 'Dry Fruits & Nuts'),
 (r'\bSUGAR\b|JAGGERY|VELLAM|\bHONEY\b|PALM SUGAR|PALM CANDY|KALKAND|\bMISRI\b|NATTU SAKKARAI', 'Sweeteners'),
 (r'SHARBATH|SHARBAT|\bSQUASH\b|JUICE|FRUTI|MAAZA|SLICE|PEPSI|\bCOKE\b|COCA|SPRITE|7 ?UP|MIRINDA|THUMS|FANTA|\bSODA\b|BISLERI|\bAQUA\b|MINERAL WATER|MOGU MOGU|\bPRAN\b|NANNARI', 'Beverages'),
 (r'\bSALT\b|PICKLE|\bPIC\b|\bSAUCE\b|KETCHUP|VINEGAR|SOYA SAUCE|TAMARIND|\bPULI\b|\bIMLI\b|\bJAM\b|GARLIC PIC|GINGER SWEET', 'Salt & Condiments'),
 (r'PEPPER|\bJEERA\b|ELACHI|CARDAMOM|\bCHILL?Y\b|CHILLI|CORIANDER|DHANIA|TURMERIC|MUSTARD|KADUGU|\bSOMBU\b|FENNEL|FENUGREEK|VENDHAYAM|CINNAMON|\bPATTAI\b|\bCLOVE\b|LAVANGAM|ASAFOET|\bHING\b|PERUNGAYAM|MASALA|SAMBAR|RASAM|GARAM|BRIYANI|BIRYANI|KULAMBU|\bAACHI\b|SAKTHI|\bMTR\b|EASTERN|STEMLESS|PULIYOG|VANGI|BISIBELA|BISI BELA', 'Spices & Masalas'),
 (r'CHOCOLATE|DAIRY MILK|KIT ?KAT|\bMUNCH\b|MILKY ?BAR|FIVE STAR|5 ?STAR|\bPERK\b|\bGEMS\b|ECLAIR|TOFFEE|\bCANDY\b|LOLLIPOP|MENTOS|\bPOLO\b|CHOCO ?PIE|BURFI|NUTTY|GULAB JAM', 'Chocolates & Candies'),
 (r'\bWAFER\b|BISCUIT|\bRUSK\b|\bBREAD\b|\bBUN\b|\bCAKE\b|COOKIE|MURUKKU|MIXTURE|CHIPS|KURKURE|KUR KURE|\bLAYS\b|\bSNACK|APPALAM|PAPAD|PAPPAD|VADAGAM|VATHAL|NIPPAT|GOOD ?DAY|\bOREO\b|HIDE.*SEEK|\bMARIE\b|BOURBON|\bTIGER\b|KRACK|POPCORN|POP CORN|ACT II|\bPUFF\b|SAMOSA|BHUJIA|\bSEV\b|NAMKEEN|HALDIRAM|MUSELI|MUESLI|OATS|OAT MASALA', 'Bakery & Snacks'),
 (r'\bEGG\b|\bTOMATO\b|\bONION\b|\bPOTATO\b|VEGETABLE|\bBANANA\b|\bCARROT\b|GREENS\b|FRESH COCONUT', 'Vegetables & Fruits'),
 (r'BATTERY|\bBULB\b|\bLED\b|CHARGER|\bCABLE\b|EARPHONE|HEADPHONE|\bUSB\b|ADAPTER|TORCH|EXTENSION|\bPLUG\b|SOCKET', 'Electronics'),
 (r'\bCUP\b|PLATE|SPOON|TUMBLER|VESSEL|\bVES\b|KADAI|\bSTEEL\b|\bALU\b|ALUMINIUM|CONTAINER|LUNCH BOX|LUNCH BAG|\bBOTTLE\b|FLASK|KNIFE|CHOPPING|FOIL|CLING|\bMUG\b|STRAINER|LADLE|\bTAWA\b|COOKER|STEAMER|PRESTIGE|PIGEON|BUTTERFLY|INOX|PRESSURE', 'Kitchen Accessories'),
]
RULES = [(re.compile(p), c) for p, c in RULES]

def classify(name):
    u = name.upper()
    for rx, cat in RULES:
        if rx.search(u):
            return cat
    return None

def slugify(s):
    s = s.lower()
    s = re.sub(r'&', ' ', s)
    s = re.sub(r'[^a-z0-9]+', '-', s).strip('-')
    return s or 'item'

def titlecase(s):
    return ' '.join(w.capitalize() if not w.isupper() or len(w) > 3 else w.title() for w in s.split())


def load_shop():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    shop = {}
    def add(name, qty, sales):
        if not name: return
        name = str(name).strip()
        if not name: return
        d = shop.setdefault(name, {'qty': 0.0, 'sales': 0.0})
        try: d['qty'] += float(qty or 0)
        except (ValueError, TypeError): pass
        try: d['sales'] += float(sales or 0)
        except (ValueError, TypeError): pass
    for r in list(wb['Normal Products (0-50%)'].iter_rows(values_only=True))[2:]:
        add(r[0], r[1], r[2])
    for r in list(wb['All Excluded Products'].iter_rows(values_only=True))[2:]:
        add(r[0], r[1], r[2])
    return shop


async def main():
    commit = '--commit' in sys.argv
    shop = load_shop()

    # existing catalogue for dup detection + max id + existing slugs
    existing_sigs = {}
    existing_slugs = set()
    maxid = 0
    async for p in products_collection.find({}, {'name': 1, 'product_id': 1, 'slug': 1}):
        sig = frozenset(t for t in toks(p.get('name', '')) if t != 'DDS')
        existing_sigs[sig] = True
        if p.get('slug'): existing_slugs.add(p['slug'])
        m = re.match(r'DHAN(\d+)', p.get('product_id', '') or '')
        if m: maxid = max(maxid, int(m.group(1)))
    sig_list = list(existing_sigs)

    def is_dup(name):
        s = frozenset(t for t in toks(name) if t != 'DDS')
        if not s: return False
        if s in existing_sigs: return True
        for a in sig_list:
            if len(s & a) >= 2 and len(s & a) / len(s) >= 0.6:
                return True
        return False

    docs = []
    used_slugs = set(existing_slugs)
    nxt = maxid
    for name, d in shop.items():
        cat = classify(name)
        if not cat:
            continue
        if is_dup(name):
            continue
        qty, sales = d['qty'], d['sales']
        price = round(sales / qty) if qty > 0 and sales > 0 else 0
        nxt += 1
        base_slug = slugify(name)
        slug = base_slug
        k = 2
        while slug in used_slugs:
            slug = f"{base_slug}-{k}"; k += 1
        used_slugs.add(slug)
        now = datetime.now(timezone.utc).isoformat()
        docs.append({
            'product_id': f"DHAN{nxt:04d}",
            'name': titlecase(name),
            'slug': slug,
            'brand': '',
            'category': cat,
            'category_slug': slugify(cat),
            'unit': '1 pc',
            'price': price,
            'base_price': price,
            'gst': 0,
            'in_stock': False,
            'stock_quantity': 0,
            'image_url': '',
            'description': None,
            'tags': [cat.lower()],
            'is_active': False,
            'source': 'sales_import_may2026',
            'created_at': now,
            'updated_at': now,
        })

    from collections import Counter
    cc = Counter(x['category'] for x in docs)
    print(f"High-confidence candidates to insert: {len(docs)}")
    for c, n in cc.most_common():
        print(f"  {c:<25} {n}")
    print("\nSamples:")
    for x in docs[:8]:
        print(f"  [{x['product_id']}] {x['name'][:38]:<38} {x['category']:<20} Rs{x['price']}")
    n0 = sum(1 for x in docs if x['price'] == 0)
    print(f"\nProducts with price 0 (no usable sales/qty): {n0}")

    if not commit:
        print("\n(DRY RUN — no writes. Re-run with --commit to insert.)")
        return
    res = await products_collection.insert_many(docs)
    print(f"\nINSERTED {len(res.inserted_ids)} products as inactive drafts (is_active=False).")

asyncio.run(main())
